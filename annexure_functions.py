from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import io

# =====================================================
#                   Annexure – 1
# =====================================================
def annexure1_generate_excel_bytes(df: pd.DataFrame) -> BytesIO:

    required_cols = [
        'Branch', 'Vendor Name', 'Product Department',
        'MRP', 'Sold Qty', 'Sold Value', 'Total LandedCost'
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    df['Profit'] = df['Sold Value'] - df['Total LandedCost']

    summary = (
        df.groupby(['Branch', 'Vendor Name', 'Product Department'], as_index=False)
        .agg({
            'MRP': 'sum',
            'Sold Qty': 'sum',
            'Sold Value': 'sum',
            'Total LandedCost': 'sum',
            'Profit': 'sum'
        })
    )

    summary['Margin (%)'] = ((summary['Profit'] / summary['Sold Value']) * 100).round(2)

    branch_sales = summary.groupby('Branch')['Sold Value'].sum().sort_values(ascending=False)
    summary['Branch'] = pd.Categorical(
        summary['Branch'],
        categories=branch_sales.index.tolist(),
        ordered=True
    )
    summary.sort_values(by=['Branch', 'Sold Value'], ascending=[True, False], inplace=True)

    wb = Workbook()
    std = wb.active
    wb.remove(std)

    out = BytesIO()

    for dept, dept_df in summary.groupby("Product Department"):
        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {str(dept).upper()}",
            "Annexure - I",
            "Vendor Wise Margin",
            "(Amount in Rs.)"
        ]

        for row_idx, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            cell = ws.cell(row=row_idx, column=1, value=text)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True, size=12)

        start_row = len(headers) + 2

        table_cols = [
            'Branch', 'Vendor Name', 'Product Department',
            'MRP', 'Sold Qty', 'Sold Value',
            'Total LandedCost', 'Profit', 'Margin (%)'
        ]

        for col_idx, col_name in enumerate(table_cols, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = Font(bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(dept_df[table_cols], index=False, header=False),
                                    start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    wb.save(out)
    out.seek(0)
    return out


# =====================================================
#                   Annexure – 2
# =====================================================
def annexure2_generate_excel_bytes(df):

    df.columns = df.columns.str.strip()

    required_cols = [
        "Branch", "Brand", "Product Department",
        "MRP", "Sold Qty", "Sold Value", "Total LandedCost", "Profit"
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise Exception(f"Missing required columns: {missing}")

    if "Profit" not in df.columns:
        df["Profit"] = df["Sold Value"] - df["Total LandedCost"]

    df["Margin %"] = (df["Profit"] / df["Sold Value"]) * 100
    df["Margin %"] = df["Margin %"].round(2)

    summary = (
        df.groupby(["Branch", "Brand", "Product Department"], as_index=False)
        .agg({
            "MRP": "sum",
            "Sold Qty": "sum",
            "Sold Value": "sum",
            "Total LandedCost": "sum",
            "Profit": "sum",
            "Margin %": "mean"
        })
    )

    summary = summary.sort_values(by=["Branch", "Sold Value"], ascending=[True, False])

    wb = Workbook()
    wb.remove(wb.active)

    for dept, data in summary.groupby("Product Department"):

        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 to 31-OCT-2025",
            f"DEPARTMENT - {dept}",
            "Annexure-II",
            "Brand Wise Margin",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")

        for row in dataframe_to_rows(data, index=False, header=True):
            ws.append(row)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =====================================================
#                   Annexure – 3
# =====================================================
def annexure3_generate_excel_bytes(df):

    df.columns = df.columns.str.strip()

    required_cols = ["Branch", "Brand", "Product Department", "Sold Qty", "Sold Value"]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    summary = (
        df.groupby(["Branch", "Brand", "Product Department"], as_index=False)
        .agg({"Sold Qty": "sum", "Sold Value": "sum"})
    )

    wb = Workbook()
    wb.remove(wb.active)

    for dept, data in summary.groupby("Product Department"):

        branch_order = (
            data.groupby("Branch")["Sold Value"]
            .sum()
            .sort_values(ascending=False)
            .index
            .tolist()
        )

        data["Branch"] = pd.Categorical(data["Branch"], categories=branch_order, ordered=True)
        data = data.sort_values(["Branch", "Sold Value"], ascending=[True, False]).reset_index(drop=True)

        data["Rank"] = (
            data.groupby("Branch")["Sold Value"]
            .rank(method="dense", ascending=False)
            .astype(int)
        )

        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 to 31-OCT-2025",
            f"DEPARTMENT - {dept}",
            "Annexure - III",
            "Brand Wise Sales Quantity & Value",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")

        for row in dataframe_to_rows(data, index=False, header=True):
            ws.append(row)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col if c.value not in (None, "")), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =====================================================
#                   Annexure – 4
# =====================================================
def annexure4_generate_excel_bytes(df):

    df.columns = df.columns.str.strip().str.lower()

    required_cols = [
        "branch", "brand", "product category",
        "product department", "sold qty", "sold value", "profit"
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns in Excel: {missing}")

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine="openpyxl")

    for dept, df_dept in df.groupby("product department"):

        summary = (
            df_dept.groupby(["branch", "brand", "product category"], as_index=False)
            .agg({
                "sold qty": "sum",
                "sold value": "sum",
                "profit": "sum"
            })
        )

        branch_totals = summary.groupby("branch", as_index=False)["sold value"] \
            .sum().rename(columns={"sold value": "branch_total"})

        brand_totals = summary.groupby(["branch", "brand"], as_index=False)["sold value"] \
            .sum().rename(columns={"sold value": "brand_total"})

        summary = summary.merge(branch_totals, on="branch", how="left")
        summary = summary.merge(brand_totals, on=["branch", "brand"], how="left")

        summary["Rank"] = (
            summary.groupby(["branch", "brand"])["sold value"]
            .rank(method="dense", ascending=False)
            .astype(int)
        )

        summary = summary.sort_values(
            by=["branch_total", "branch", "brand_total", "brand", "sold value"],
            ascending=[False, True, False, True, False]
        ).reset_index(drop=True)

        summary = summary[[
            "branch", "brand", "product category",
            "sold qty", "sold value", "Rank"
        ]]

        safe_sheet = str(dept).replace("/", "_")[:31]
        summary.to_excel(writer, index=False, sheet_name=safe_sheet, startrow=7)

        ws = writer.book[safe_sheet]

        header_texts = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - IV",
            "Product Wise Sales Quantity And Value",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(header_texts, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            cell = ws.cell(row=i, column=1)
            cell.value = text
            cell.font = Font(name="Calibri", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    writer.close()
    output.seek(0)
    return output

# =====================================================
#                   Annexure – 5
# =====================================================
def annexure5_generate_excel_bytes(df):

    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    # ----------------------------
    # Step 1: Clean columns
    # ----------------------------
    df.columns = df.columns.str.strip().str.lower()

    # ----------------------------
    # Step 2: Validate required columns
    # ----------------------------
    required_cols = [
        "branch", "product category", "product department",
        "sold qty", "sold value"
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns: {missing}")

    # ----------------------------
    # Step 3: Create Workbook
    # ----------------------------
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ----------------------------
    # Step 4: Loop Department-wise
    # ----------------------------
    for dept, df_dept in df.groupby("product department"):

        # --- 4.1: Aggregate ---
        summary = (
            df_dept.groupby(["branch", "product category"], as_index=False)
            .agg({
                "sold qty": "sum",
                "sold value": "sum"
            })
        )

        # --- 4.2: Branch-wise total ---
        branch_total = (
            summary.groupby("branch", as_index=False)["sold value"]
            .sum()
            .rename(columns={"sold value": "branch_total"})
        )

        summary = summary.merge(branch_total, on="branch", how="left")

        # --- 4.3: % Contribution ---
        summary["%Contribution"] = (
            summary["sold value"] / summary["branch_total"] * 100
        ).map(lambda x: f"{x:.2f}%")

        # --- 4.4: Sorting ---
        summary = summary.sort_values(
            by=["branch_total", "branch", "sold value"],
            ascending=[False, True, False]
        ).reset_index(drop=True)

        # --- 4.5: Final Column Order ---
        summary = summary[
            ["branch", "product category", "sold qty", "sold value", "%Contribution"]
        ]

        # ----------------------------
        # Create Sheet
        # ----------------------------
        sheet_name = str(dept).replace("/", "_").replace("\\", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # ----------------------------
        # Header Section
        # ----------------------------
        header_texts = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - V",
            "Product Category Contribution - All Branches",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(header_texts, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
            cell = ws.cell(row=i, column=1)
            cell.value = text
            cell.font = Font(name="Calibri", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ----------------------------
        # Data Table Start Row
        # ----------------------------
        start_row = 8

        # --- Write header row ---
        for col_idx, col_name in enumerate(summary.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        # --- Write data rows ---
        for row_idx, row in enumerate(summary.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # --- Auto column width ---
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    # ----------------------------
    # RETURN EXCEL BYTES
    # ----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output



# Annexure_function.py with Annexure 6 added

# Paste your existing Annexure 1–5 code above this section

# =====================================================
#                  Annexure – 6
# =====================================================
def annexure6_generate_excel_bytes(df):
    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # Normalize column names
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    # Helper to find column
    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    item_code_col = find_col(["item code", "itemcode"])
    product_name_col = find_col(["product name", "productname"])
    branch_col = find_col(["branch"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold qty", "soldqty"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    landed_col = find_col(["total landedcost", "total landed cost"])
    profit_col = find_col(["profit"])
    dept_col = find_col(["product department", "productdepartment"])

    if not all([sold_qty_col, sold_value_col, landed_col, profit_col, branch_col]):
        raise KeyError("Missing one or more required columns: Sold Qty, Sold Value, Landed Cost, Profit, Branch")

    if dept_col is None:
        df["product department"] = "All"
        dept_col = "product department"

    # Convert numeric fields
    for col in [sold_qty_col, sold_value_col, landed_col, profit_col, mrp_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Derived metrics
    df["Sales per Qty"] = np.where(df[sold_qty_col] != 0, df[sold_value_col] / df[sold_qty_col], 0)
    df["L Cost per Qty"] = np.where(df[sold_qty_col] != 0, df[landed_col] / df[sold_qty_col], 0)
    df["Profit % on Sales"] = np.where(df[sold_value_col] != 0, (df[profit_col] / df[sold_value_col]) * 100, 0)

    # Filter Profit > 0 and Profit% < 10
    filtered = df[(df[profit_col] > 0) & (df["Profit % on Sales"] < 10)].copy()

    if filtered.empty:
        raise ValueError("No Products Found with Profit Below 10% and Greater Than 0.")

    # Prepare output workbook (multi-sheet)
    wb = Workbook()
    wb.remove(wb.active)

    for dept, g in filtered.groupby(dept_col):
        g = g.reset_index(drop=True)
        g["S.No"] = range(1, len(g) + 1)

        final = pd.DataFrame({
            "S.No": g["S.No"],
            "Item Code": g.get(item_code_col, ""),
            "Product Name": g.get(product_name_col, ""),
            "Branch": g[branch_col],
            "MRP": g[mrp_col].round(2),
            "Sales per Qty": g["Sales per Qty"].round(2),
            "L Cost per Qty": g["L Cost per Qty"].round(2),
            "Sold Qty": g[sold_qty_col].round(2),
            "Sold Value": g[sold_value_col].round(2),
            "Total Landed Cost": g[landed_col].round(2),
            "Profit": g[profit_col].round(2),
            "Profit % on Sales": g["Profit % on Sales"].round(2)
        })

        # Sheet creation
        safe_name = str(dept).replace("/", "_").replace("\\", "_")[:31]
        ws = wb.create_sheet(title=safe_name)

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - VI",
            "List of Products Sold at a Profit Below 10%",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
            c = ws.cell(row=i, column=1, value=text)
            c.font = Font(bold=True, size=12)
            c.alignment = Alignment(horizontal="center", vertical="center")

        start_row = len(headers) + 2

        # Column headers
        for j, col in enumerate(final.columns, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto column width
        for i, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
            ws.column_dimensions[col_letter].width = max_len + 3

    # Return file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def annexure8_generate_excel_bytes(df):
    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # Normalize column names
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    # Helper to find column names
    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    # Identify required columns
    name_col = find_col(["name"])
    location_col = find_col(["location"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold quantity", "sold qty", "soldqty"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    net_value_col = find_col(["net value", "netvalue"])
    product_code_col = find_col(["product code", "productcode"])
    dept_col = find_col(["department", "product department", "productdepartment"])

    # Validate required columns
    req = [sold_qty_col, sold_value_col, net_value_col]
    if not all(req):
        raise KeyError("Missing required columns: Sold Quantity, Sold Value, Net Value")

    # If no department column -> assign ALL
    if dept_col is None:
        df["department"] = "ALL"
        dept_col = "department"

    # Convert numerics
    for col in [mrp_col, sold_qty_col, sold_value_col, net_value_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Filter Sold Qty > 0
    df = df[df[sold_qty_col] > 0].copy()

    # Calculations
    df["Sales per Qty"] = df[sold_value_col] / df[sold_qty_col]
    df["Lcost per Qty"] = df[net_value_col] / df[sold_qty_col]

    # Filter: Selling price < Landed cost
    df = df[df["Sales per Qty"] < df["Lcost per Qty"]].copy()

    df["Total Sales"] = df[sold_value_col]
    df["Total Landed Cost"] = df[net_value_col]
    df["Total Loss"] = df["Total Sales"] - df["Total Landed Cost"]

    # Round
    round_cols = ["Sales per Qty", "Lcost per Qty", "Total Sales", "Total Landed Cost", "Total Loss"]
    df[round_cols] = df[round_cols].round(2)

    # Prepare output workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Group by department
    for dept, g in df.groupby(dept_col):
        g = g.reset_index(drop=True)
        g["S.No"] = range(1, len(g) + 1)

        final = pd.DataFrame({
            "S.No": g["S.No"],
            "Name": g.get(name_col, ""),
            "Location": g.get(location_col, ""),
            "MRP": g.get(mrp_col, "").round(2),
            "Sold Quantity": g[sold_qty_col],
            "Sales per Qty": g["Sales per Qty"],
            "Lcost per Qty": g["Lcost per Qty"],
            "Total Sales": g["Total Sales"],
            "Total Landed Cost": g["Total Landed Cost"],
            "Total Loss": g["Total Loss"],
            "Product Code": g.get(product_code_col, "")
        })

        # Sheet name safe
        safe_name = str(dept).replace("/", "_")[:31]
        ws = wb.create_sheet(title=safe_name)

        # Header lines
        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - VIII",
            "Selling Price Less Than Purchase Cost",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        start_row = len(headers) + 2

        # Column headers
        for j, col in enumerate(final.columns, start=1):
            c = ws.cell(row=start_row, column=j, value=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto column width
        for i, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
            ws.column_dimensions[col_letter].width = max_len + 3

    # Save output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def annexure9_generate_excel_bytes(df):
    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # Normalize column names
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    # Helper to detect columns
    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    item_code_col = find_col(["item code", "itemcode", "product code"])
    product_name_col = find_col(["product name", "productname", "name"])
    branch_col = find_col(["branch"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold qty", "soldquantity", "sold quantity"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    landed_col = find_col(["net value", "netvalue", "total landedcost", "total landed cost"])
    dept_col = find_col(["department", "product department", "productdepartment"])

    # REQUIRED COLUMNS CHECK
    required = [sold_qty_col, sold_value_col, landed_col]
    if not all(required):
        raise KeyError("Missing required columns: Sold Quantity, Sold Value, Landed Cost")

    if dept_col is None:
        df["department"] = "All"
        dept_col = "department"

    df = df[df[sold_qty_col] > 0].copy()

    # Convert to numeric
    for col in [sold_qty_col, sold_value_col, landed_col, mrp_col]:
        if col:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # DERIVED METRICS
    df["Sales per Qty"] = np.where(df[sold_qty_col] != 0, df[sold_value_col] / df[sold_qty_col], 0)
    df["L Cost per Qty"] = np.where(df[sold_qty_col] != 0, df[landed_col] / df[sold_qty_col], 0)

    # CONDITION FOR ANNEXURE 9 — ZERO PROFIT / ZERO LOSS
    filtered = df[np.isclose(df[sold_value_col], df[landed_col], atol=0.01)].copy()

    if filtered.empty:
        raise ValueError("No products found where Sold Value = Total Landed Cost.")

    # PREPARE WORKBOOK
    wb = Workbook()
    wb.remove(wb.active)

    for dept, g in filtered.groupby(dept_col):
        g = g.reset_index(drop=True)
        g["S.No"] = range(1, len(g) + 1)

        # Prepare final sheet data
        final = pd.DataFrame({
            "S.No": g["S.No"],
            "Item Code": g.get(item_code_col, ""),
            "Product Name": g.get(product_name_col, ""),
            "Branch": g.get(branch_col, ""),
            "MRP": g.get(mrp_col, 0).round(2),
            "Sold Quantity": g[sold_qty_col].round(2),
            "Sales per Qty": g["Sales per Qty"].round(2),
            "L Cost per Qty": g["L Cost per Qty"].round(2),
            "Sold Value": g[sold_value_col].round(2),
            "Total Landed Cost": g[landed_col].round(2),
            "Difference (Sales - Landed)": (g[sold_value_col] - g[landed_col]).round(2)
        })

        # Sheet creation
        safe_name = str(dept).replace("/", "_").replace("\\", "_")[:31]
        ws = wb.create_sheet(title=safe_name)

        # HEADINGS
        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - IX",
            "List of Products Checked & Matched (Neither Profit Nor Loss)",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
            c = ws.cell(row=i, column=1, value=text)
            c.font = Font(bold=True, size=12)
            c.alignment = Alignment(horizontal="center", vertical="center")

        start_row = len(headers) + 2

        # Column headers
        for j, col in enumerate(final.columns, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto column width
        for i, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
            ws.column_dimensions[col_letter].width = max_len + 3

    # Return file as bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output




def annexure10_generate_excel_bytes(df):
    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    # Keep only rows where Sold Quantity > 0
    df = df[df["Sold Quantity"] > 0].copy()

    # Calculations
    df["Sales per Qty"] = df["Sold Value"] / df["Sold Quantity"]
    df["Lcost per Qty"] = df["Net Value"] / df["Sold Quantity"]
    df["Vendor Margin %"] = ((df["MRP"] - df["Lcost per Qty"]) * 100) / df["MRP"]
    df["Profit Margin %"] = ((df["Sales per Qty"] - df["Lcost per Qty"]) * 100) / df["Sales per Qty"]

    # Profit Amount per unit
    df["Profit Amount"] = df["Sales per Qty"] - df["Lcost per Qty"]

    # Rounding
    df["Sales per Qty"] = df["Sales per Qty"].round(2)
    df["Lcost per Qty"] = df["Lcost per Qty"].round(2)
    df["Vendor Margin %"] = df["Vendor Margin %"].round(2)
    df["Profit Margin %"] = df["Profit Margin %"].round(2)
    df["Profit Amount"] = df["Profit Amount"].round(2)

    # ===============================
    #       NEW FILTER CONDITIONS
    # ===============================
    df = df[
        (df["Profit Amount"] > 0) &
        (df["Vendor Margin %"] >= 40) &
        (df["Profit Margin %"] <= 40) &
        (df["Profit Margin %"] > 0) 
    ]

    # Create workbook with multiple sheets per department
    wb = Workbook()
    wb.remove(wb.active)

    for dept, g in df.groupby("Department"):
        g = g.reset_index(drop=True)
        g["S.No"] = range(1, len(g) + 1)

        final_cols = [
            "Name", "Location", "MRP", "Sold Quantity",
            "Sales per Qty", "Lcost per Qty", "Profit Amount",
            "Vendor Margin %", "Profit Margin %",
            "Product Code"
        ]
        final = g[final_cols]

        ws = wb.create_sheet(title=str(dept)[:31])

        # Add headings
        headings = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept}",
            "Annexure - X",
            "High Vendor Margin Less Profit Margin",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headings, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Column headers
        start_row = len(headings) + 1
        for j, col in enumerate(final.columns, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def annexure11_generate_excel_bytes(closing_df, sales_df, ibts_df):
    import pandas as pd
    from io import BytesIO
    import numpy as np

    # --- CLEAN COLUMNS ---
    def clean_cols(df):
        df.columns = df.columns.map(lambda x: str(x).replace("\u00a0", "").strip())
        return df

    closing_df = clean_cols(closing_df)
    sales_df   = clean_cols(sales_df)
    ibts_df    = clean_cols(ibts_df)

    # --- RENAME PRODUCT ATTRIBUTE ---
    sales_df.rename(columns={"Product Attribute Id": "Product Attribute"}, inplace=True)
    ibts_df.rename(columns={"Product Attribute Id": "Product Attribute"}, inplace=True)

    if "Product Attribute" not in closing_df.columns:
        raise Exception("❌ Product Attribute missing in Closing Stock")

    # --- FILTER IBTS LBTS ONLY ---
    if "Doc No" not in ibts_df.columns:
        raise Exception("❌ Doc No missing in IBTS file")

    ibts_df["Doc No"] = ibts_df["Doc No"].astype(str)
    ibts_df = ibts_df[ibts_df["Doc No"].str.contains("LBTS", case=False, na=False)]
    ibts_df = ibts_df[~ibts_df["Doc No"].str.contains("LBTR", case=False, na=False)]

    # --- NORMALIZE ATTRIBUTE ---
    def fix(df):
        df["Product Attribute"] = df["Product Attribute"].astype(str).str.replace(".0", "", regex=False)
        df["Product Attribute"] = pd.to_numeric(df["Product Attribute"], errors="coerce")
        df.dropna(subset=["Product Attribute"], inplace=True)

    fix(closing_df)
    fix(sales_df)
    fix(ibts_df)

    sales_set = set(sales_df["Product Attribute"])
    ibts_set  = set(ibts_df["Product Attribute"])

    # --- GET NON MOVEMENT ---
    non_move = closing_df[
        (~closing_df["Product Attribute"].isin(sales_set)) &
        (~closing_df["Product Attribute"].isin(ibts_set))
    ].copy()

    # --- FILTER TAXABLE ---
    non_move["Taxable Amount"] = pd.to_numeric(non_move["Taxable Amount"], errors="coerce")
    non_move = non_move[non_move["Taxable Amount"] != 0]

    # --- QTY COLUMN ---
    qty_candidates = ["Onhand Qty", "On Hand Qty", "Inhand Qty", "Stock Qty"]
    qty_col = next((c for c in qty_candidates if c in non_move.columns), None)

    if not qty_col:
        raise Exception("Onhand Qty column not found")

    non_move[qty_col] = pd.to_numeric(non_move[qty_col], errors="coerce")
    non_move = non_move[non_move[qty_col] > 0]

    # --- DEPARTMENT ---
    if "Department" not in non_move.columns:
        non_move["Department"] = "Unknown"

    # --- FINAL COLUMNS ---
    final_cols = [
        "Name", "Department", "Category", "Sub Category", "Brand",
        "Vendor", "Product Design", qty_col, "Value By MRP",
        "Taxable Amount", "Product Attribute", "Tax Amount"
    ]
    final_cols = [c for c in final_cols if c in non_move.columns]

    # --- CREATE EXCEL ---
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        for dept, df in non_move.groupby("Department"):
            safe = str(dept).replace("/", "_")[:31]
            sheet = writer.book.add_worksheet(safe)
            writer.sheets[safe] = sheet

            headers = [
                "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
                "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
                f"DEPARTMENT - {dept}",
                "Annexure - XI",
                "Non Movement of Stock for the period of 3 months",
                "(Amount in Rs.)"
            ]

            for i, text in enumerate(headers):
                sheet.merge_range(i, 0, i, len(final_cols)-1, text)

            df[final_cols].to_excel(writer, sheet_name=safe, startrow=len(headers)+1, index=False)

    output.seek(0)
    return output


